"""Inspect the generated Excel model for data issues and formatting problems.

Run with the project venv: python tools/inspect_excel.py output/AAPL_model.xlsx
"""
import sys
from pathlib import Path
import pandas as pd
import math
from openpyxl import load_workbook


def check_dataframe(df: pd.DataFrame, name: str):
    out = []
    out.append(f"Sheet '{name}': shape={df.shape}")
    if df.empty:
        out.append("  - WARNING: sheet is empty")
        return out

    # assume first column is labels for financial statements
    if df.shape[1] >= 2:
        numeric = df.iloc[:, 1:]
    else:
        numeric = df.select_dtypes(include=["number"])

    # coerce any convertible columns to numeric (strings like '1,234' or formatted numbers)
    if not numeric.empty:
        numeric = numeric.apply(lambda col: pd.to_numeric(col.astype(str).str.replace(',',''), errors='coerce'))

    if numeric.empty:
        out.append("  - No numeric data detected (may be overview or notes)")
        return out

    # NaN summary
    nan_pct = numeric.isna().mean().mean()
    out.append(f"  - Numeric columns: {numeric.shape[1]}, overall NaN%={nan_pct:.1%}")

    # magnitude checks
    # compute max absolute value among numeric cells (ignore non-numeric coerced to NaN)
    max_abs = numeric.abs().max().max()
    out.append(f"  - Max absolute value among numeric cells: {max_abs}")
    if max_abs > 1e13:
        out.append("  - FLAG: Extremely large values (>1e13) found")

    # period-to-period change check: per-row relative change
    try:
        changes = []
        for idx, row in numeric.iterrows():
            vals = row.dropna().values
            if len(vals) >= 2 and all((isinstance(x, (int, float)) and not math.isinf(x)) for x in vals):
                # compute max relative change
                rel = max(abs((vals[i+1] - vals[i]) / (vals[i] if vals[i] != 0 else 1)) for i in range(len(vals)-1))
                changes.append(rel)
        if changes:
            max_rel = max(changes)
            out.append(f"  - Max period-to-period relative change among rows: {max_rel:.1%}")
            if max_rel > 5.0:
                out.append("  - FLAG: Very large single-period change (>500%) found")
    except Exception:
        out.append("  - Could not compute period-to-period changes (non-numeric layout)")

    return out


def inspect_formats(path: Path, sheet_name: str, sample_rows=5):
    wb = load_workbook(path, data_only=False)
    if sheet_name not in wb.sheetnames:
        return [f"Sheet '{sheet_name}' not found for format inspection"]
    ws = wb[sheet_name]
    rows = []
    rows.append(f"Number formats sample for sheet '{sheet_name}':")
    max_row = min(sample_rows, ws.max_row)
    max_col = min(10, ws.max_column)
    for r in range(1, max_row+1):
        rowfmts = []
        for c in range(1, max_col+1):
            cell = ws.cell(row=r, column=c)
            fmt = cell.number_format
            # simplify empty
            if not fmt:
                fmt = 'GENERAL'
            rowfmts.append(fmt)
        rows.append('  ' + ', '.join(rowfmts))
    return rows


def main():
    if len(sys.argv) < 2:
        print("Usage: python inspect_excel.py <path_to_xlsx>")
        raise SystemExit(1)
    path = Path(sys.argv[1])
    if not path.exists():
        print(f"File not found: {path}")
        raise SystemExit(1)

    sheets = ['Overview', 'Income Statement', 'Balance Sheet', 'Cash Flow', 'Historical Prices', 'Ratios', 'Forecast']

    print(f"Inspecting: {path}\n")
    # read with pandas for data checks
    for s in sheets:
        try:
            df = pd.read_excel(path, sheet_name=s, engine='openpyxl')
        except Exception:
            print(f"- Could not read sheet '{s}' as DataFrame")
            continue
        for line in check_dataframe(df, s):
            print(line)
        print('')

    # format inspection using openpyxl
    print('\nFormat checks (sample rows):')
    for s in sheets:
        for line in inspect_formats(path, s):
            print(line)
        print('')


if __name__ == '__main__':
    main()
